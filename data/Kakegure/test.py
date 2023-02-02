from riotwatcher import LolWatcher, ApiError
import pandas as pds
import numpy as num
from pathlib import Path
from os.path import exists
#import xlsxwriter
import openpyxl
import time

# global variables
api_key = 'RGAPI-769ca7e5-5def-4e0a-8b03-79cc2a142125'

# List of Summoners (Max 5)
summoners = [ 'UnawareIdiot99' , 'Spykid810' , 'Kaminaryuu' ,
                'EáZy' , 'Zond Lilac' , 'Physıcal' ]

# Create a blank Game_Data database
file_name = 'Game_Data.xlsx'
# Create a new workbook
workbook = openpyxl.Workbook()
# Save the workbook as an Excel file
workbook.save(file_name)

# Check if file exists or not
PATH = '/workspaces/codespaces-jupyter/Game_Data.xlsx'
file_exists = exists(PATH)

# Function to Analyze Game Data per Summoner Given
def game_analyzer(api_key, summonerName, queue):
    # Arg 1: API Key String, use your own personal key from Riot Dev URL
    # Arg 2: List of Summoner Names you want to retrieve the data from. 
    # Arg 3: Type of Queue
    
    # Setting Up Watcher from API
    watcher = LolWatcher(api_key)

    # Setting Region and collecting Summoner Data by Name
    my_region = 'na1'

    # Type of Queue (Solo, Flex, Custom)
    if (queue == 'solo'):
        queue = 420
    if queue == 'flex':
        queue = 440
    if queue == 'custom':
        queue = 0

    # Get a list of player names for each summoner
    summonerDict = {}

    # For each Summoner Name given on the list, we will retrieve all the latest games
    # so we can build our own database for each player
    for playerName in summonerName:
        # Get the player data from league
        playerLeagueInfo = watcher.summoner.by_name(my_region, playerName)
        # Return the rank status for the Summoner
        #rankedStats = watcher.league.by_summoner(my_region, playerLeagueInfo['id'])
        # Max number of Games that can be retrieved: 100
        maxNum = 100
        time.sleep(5)
        matches = watcher.match.matchlist_by_puuid(my_region, playerLeagueInfo['puuid'],0,maxNum,queue)

        # Now we will collect the data for the last 50 gamesplayerDictplayerDict
        games = 0
        # new counter
        new = 1
        # We will make a temp list for each player
        playerList = []
        # Get a local variable to make a list of all games
        while games < maxNum:
            # Get Data for each player
            singleGame = single_game(my_region, matches[games], playerName)
            # local list of the game data
            list = singleGame[0]
            playerList.append(singleGame[0])

            # Get the Data Frame output from Single Game
            gameDF = singleGame[1]

            # Check if file already exists in the server
            if file_exists is True:
                # Check if file is empty
                reader = pds.read_excel(file_name)
                r = reader.empty
                if r is True:
                    # Write initial data
                    saver( gameDF , playerName , file_name )
                else:
                    # Retrieve the column/header data which are the keys of the dictionary
                    # check if list is empty, if it is, skip
                    if list == []:
                        # Dont save anything
                        continue
                    else:
                        # Get the keys for the list
                        keys = list[0].keys()
                        
                    # print(keys)
                        # get number of iterations
                        n = 1
                        for k in keys:
                            d = list[0].get(k)
                            
                            # Opening the Excel File
                            wb = openpyxl.load_workbook('Game_Data.xlsx')

                            # Check if each player has its own sheet, otherwise, create one
                            if playerName in wb:
                                # create a sheet variable
                                playerSheet = wb[playerName]
                                # First row of each sheet is the same
                                playerSheet.cell( row = 1 , column = n ).value = k
                                # The following rows, are the new data to be added
                                playerSheet.cell( row = new , column = n ).value = d
                                wb.save(file_name)
                            else:
                                playerSheet = wb.create_sheet(playerName)
                                # First row of each sheet is the same
                                playerSheet.cell( row = 1 , column = n ).value = k
                                # The following rows, are the new data to be added
                                playerSheet.cell( row = new , column = n ).value = d
                                wb.save(file_name)
                            n += 1

            # Update the counter for the loop
            games += 1
            new += 1



# Function to retrieve data for single game
def single_game(my_region, lastMatch, summonerGiven):
    # Setting Up Watcher from API
    watcher = LolWatcher(api_key)

    # Getting the data from last match in dictionary format
    time.sleep(5)
    match_detail = watcher.match.by_id(my_region, lastMatch)

    # Removing info and metadata keys - no needed
    remove_info = match_detail.pop('metadata')
    remove_info = match_detail.pop('info')

    # Getting List of Participants
    participants = remove_info.get('participants')

    # Player and Combined Dictionaries
    playerData = []
    
    # Now to Retrieve the actual Game Data
    for row in remove_info:
        # Game Data is inside Participants Key
        if row == 'participants':
            # Local Varialbe for the new Dictionary to be traversed
            LL = remove_info[row]

            # Traversing Each Player Data (pd) on the Game        
            for pd in remove_info[row]:
                # Player Data local variable
                pdll = {}

                # For SoloQ Stats, we only want the data for the summoner which we have given
                # any random is not wanted
                # We first get the summoner name value of the game data
                summonerName = pd.get('summonerName')

                # Now we verify it against what we gave, if it is a match, we continue, if not
                # we pass
                if summonerName == summonerGiven:
                    # Some data that we want is under another dictionary called
                    # challenges. Before that iteration, we will take the data we want
                    # from the keys that are accessible from here

                    # Champion Selection, Lane and Role
                    champSelected = pd.get('championName')
                    # for Wukong
                    if champSelected == 'MonkeyKing':
                        champSelected = 'Wukong'
                    if pd.get('teamPosition') == 'UTILITY':
                        lane = 'SUPPORT'
                    else:
                        lane = pd.get('teamPosition')

                    # Champion Data
                    champLevel = pd.get('champLevel')
                    kills = pd.get('kills')
                    assists = pd.get('assists')
                    deaths = pd.get('deaths')
                    
                    # Performance Metrics In General
                    triple = pd.get('tripleKills')
                    quadra = pd.get('quadraKills')
                    penta = pd.get('pentaKills')
                    objetivesStolen = pd.get('objectivesStolen')
                    firstBlood = pd.get('firstBloodKill')

                    # Total Damage Chart
                    damageToChamp = pd.get('totalDamageDealtToChampions')
                    damageTaken = pd.get('totalDamageTaken')

                    # Gold Chart
                    totalGold = pd.get('goldEarned')

                    # Warding and vision score
                    wardPlaced = pd.get('wardsPlaced')
                    wardKilled = pd.get('wardsKilled')
                    wardBought = pd.get('visionWardsBoughtinGame')
                    vs = pd.get('visionScore')
                    
                    # Time Statistics
                    totalTimeCCing = pd.get('totalTimeCCDealt')
                    timeDead = pd.get('totalTimeSpentDead')
                    
                    # Game Results
                    timePlayed = pd.get('timePlayed')
                    result = pd.get('win')

                    # Turret Information
                    turretTakedowns = pd.get('turretTakedowns')
                    damageDealtToTurrets = pd.get('damageDealtToTurrets')
                    turretKills = pd.get('turretKills')
                    firstTurret = pd.get('firstTowerKill')
                    turretLost = pd.get('turretsLost')

                    # Surrender
                    gameSurrenders = pd.get('gameEndedInSurrender')
                    teamEarlySurrender = pd.get('teamEarlySurrendered')

                    # Total CS Data
                    laneCS = pd.get('totalMinionsKilled')
                    neutralCS = pd.get('neutralMinionsKilled')

                    # Retrieving Dict of Challenges
                    challenges = pd.get('challenges')


                    # Getting data from Challenges
                    gpm = challenges.get('goldPerMinute')
                    cward = challenges.get('controlWardsPlaced')
                    dpm = challenges.get('damagePerMinute')
                    teamDamageTakenPercentage = round(challenges.get('damageTakenOnTeamPercentage') * 100, 2)
                    kda = round(challenges.get('kda'), 2)
                    divesReceived = challenges.get('killsUnderOwnTurret')
                    # Check if Key is none, in case no KP has been identified in the game
                    if challenges.get('killParticipation') is None:
                        kp = 0
                    else:
                        kp = challenges.get('killParticipation') * 100
                    divesPerformed = challenges.get('killsNearEnemyTurret')
                    teamDamagePercentage = round(challenges.get('teamDamagePercentage') * 100, 2)
                    earlyTurrets = challenges.get('outerTurretExecutesBefore10Minutes')
                    turretPlatesTaken = challenges.get('turretPlatesTaken')
                    vpm = challenges.get('visionScorePerMinute')
                    wardTakendowns20 = challenges.get('wardTakedownsBefore20M')
                    turretWithHerald = challenges.get('turretsTakenWithRiftHerald')
                    soloKill = challenges.get('soloKills')
                    allyJG = challenges.get('alliedJungleMonsterKills')
                    enemyJG = challenges.get('enemyJungleMonsterKills')

                    # Jungle vs Lane CS                        
                    if lane == 'JUNGLE':
                        cs10 = round(challenges.get('jungleCsBefore10Minutes'))
                        goldDiff10 = challenges.get('laningPhaseGoldExpAdvantage')
                    else:
                        cs10 = round(challenges.get('laneMinionsFirst10Minutes'))
                        goldDiff10 = challenges.get('earlyLaningPhaseGoldExpAdvantage')

                    # We now calculate the correct CS for each player
                    cs = laneCS + neutralCS + round(allyJG) + round(enemyJG)
                    cs = round(cs, 2)

                    # Time Played is collected in seconds, so we need to convert into minutes
                    minutes = timePlayed / 60
                    seconds = timePlayed % 60
                    minQuotient = timePlayed // 60
                    # Therefore, CS per minute is
                    csm = cs / minutes

                    # Converting into string data
                    minSTR = str(minQuotient)
                    minSEC = str(seconds).zfill(2)
                    gameTime = minSTR + ':' + minSEC

                    # Creating Player Data Dictionary
                    pdll['Summoner Name'] = summonerName
                    pdll['Champion Selected'] = champSelected
                    pdll['Lane'] = lane
                    pdll['Champion Level'] = champLevel
                    pdll['Kills'] = kills
                    pdll['Deaths'] = deaths
                    pdll['Assists'] = assists
                    pdll['KDA'] = kda
                    pdll['Damage Per Minute'] = dpm
                    pdll['Gold Per Minute'] = gpm
                    pdll['Vision Score Per Minute'] = vpm
                    pdll['Triple Kills'] = triple
                    pdll['Quadra Kills'] = quadra
                    pdll['Penta Kills'] = penta
                    pdll['Number of Solo Kills'] = soloKill
                    pdll['Kill Participation'] = kp
                    pdll['CS@10'] = cs10
                    pdll['GD@10'] = goldDiff10
                    pdll['Damage to Champion'] = damageToChamp
                    pdll['Damage Taken'] = damageTaken
                    pdll['Team Damage Dealt Percentage'] = teamDamagePercentage
                    pdll['Team Damage Taken Percentage'] = teamDamageTakenPercentage
                    pdll['Total Gold Earned'] = totalGold
                    pdll['Control Wards Placed'] = cward
                    pdll['Wards Placed'] = wardPlaced
                    pdll['Wards Killed'] = wardKilled
                    pdll['Control Wards Bought'] = wardBought
                    pdll['Vision Score'] = vs
                    pdll['Total CC Time Dealt'] = totalTimeCCing
                    pdll['Total Time Spent Dead'] = timeDead
                    pdll['Game Ended in Surrender'] = gameSurrenders
                    pdll['Own Team Early Surrender'] = teamEarlySurrender
                    pdll['Time Played '] = gameTime
                    pdll['Win'] = result
                    pdll['Total CS'] = cs
                    pdll['CS Per Minute'] = csm
                    pdll['Turrets Taken with Herald'] = turretWithHerald
                    pdll['Turrets Takedowns'] = turretTakedowns
                    pdll['Plates Taken'] = turretPlatesTaken
                    pdll['Damage to Turrets'] = damageDealtToTurrets
                    pdll['Turrets Taken'] = turretKills
                    pdll['Turrets Lost'] = turretLost
                    pdll['Turrets Taken Before 10min'] = earlyTurrets
                    pdll['Dives Received'] = divesReceived
                    pdll['Dives Taken'] = divesPerformed
                    pdll['Objectives Stolen'] = objetivesStolen
                    pdll['First Turret'] = firstTurret
                    pdll['First Blood'] = firstBlood
                    pdll['Ward Take Downs Before 20min'] = wardTakendowns20

                    # Saving all in Game Data
                    playerData.append(pdll)
                else:
                    continue

    # Creating a Data Frame for Game Data
    playerDF = pds.DataFrame(playerData)
    # If Data Frame is empty, clear it and dont use it.
    # If there is data on it, save it
    #if playerDF.empty:
        # Clear Memory and return nothing
     #   playerData.clear()
      #  return print('Data is empty')
    #else:
        # File Name and Saving in Excel
        #file_name = 'Game_Data.xlsx'
        #playerDF.to_excel(file_name)

        # Returning Data Frame and List
    return playerData, playerDF

# Function to save data in Excel given a DataFrame
def saver(playerDF,playerName, file_name):
    # File Name and Saving in Excel   
    playerDF.to_excel(file_name, sheet_name=playerName)

game_analyzer(api_key, summoners, 'solo')
